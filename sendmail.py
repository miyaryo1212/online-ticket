import ctypes
import smtplib
import tkinter
import tkinter.filedialog
import tkinter.messagebox
from email import message
from tkinter import ttk

import openpyxl
import qrcode


def msgbox(title, content):
    root = tkinter.Tk()
    root.withdraw()
    tkinter.messagebox.showinfo(title, content)

    return None


def openfile(file_type=[("", "*")]):
    root = tkinter.Tk()
    root.withdraw()
    root.title("Excelブックを選択")
    file_path = tkinter.filedialog.askopenfilename(filetypes=file_type)

    if not file_path:
        quit()

    return file_path


def openfolder():
    root = tkinter.Tk()
    root.withdraw()
    root.title("保存先ディレクトリを選択")
    dir_path = tkinter.filedialog.askdirectory()

    if not dir_path:
        quit()

    return dir_path


def askuserinfo():
    root = tkinter.Tk()
    root.withdraw()
    root.title("ログイン情報を入力")

    def get(event):
        root.quit()

    label_account = tkinter.Label(root, text="Outlook account")
    label_account.grid(column=0, row=0, padx=20, pady=10)
    label_address = tkinter.Label(root, text="Email address")
    label_address.grid(column=0, row=1, padx=20, pady=10)
    label_password = tkinter.Label(root, text="Password")
    label_password.grid(column=0, row=2, padx=20, pady=10)

    box_account = tkinter.Entry(root, width=35)
    box_account.grid(column=1, row=0, padx=20, pady=10)
    box_address = tkinter.Entry(root, width=35)
    box_address.grid(column=1, row=1, padx=20, pady=10)
    box_password = tkinter.Entry(root, width=35, show="*")
    box_password.grid(column=1, row=2, padx=20, pady=10)

    root.bind("<Return>", get)

    root.deiconify()
    root.mainloop()
    root.withdraw()

    if not box_address.get() or not box_password.get():
        msgbox("Error", "Outlookアカウントまたはメールアドレス、パスワードが正しく入力されませんでした")
        quit()

    return box_address.get(), box_password.get()


def selectsheet(bookpath):
    workbook = openpyxl.load_workbook(bookpath)
    sheets = workbook.sheetnames

    root = tkinter.Tk()
    root.withdraw()
    root.title("ワークシートを選択")

    def get(event):
        root.quit()

    label = tkinter.Label(root, text="Sheet")
    label.grid(column=0, row=0, padx=20, pady=10)
    combo = ttk.Combobox(root, values=sheets, justify="center")
    combo.grid(column=1, row=0, padx=20, pady=10)

    root.bind("<Return>", get)

    root.deiconify()
    root.mainloop()
    root.withdraw()

    if not combo.get():
        msgbox("Error", "シートが正しく選択されませんでした")
        quit()

    return combo.get()


def askqty():
    root = tkinter.Tk()
    root.withdraw()
    root.title("")

    def get(event):
        root.quit()

    label_start = tkinter.Label(root, text="Start")
    label_start.grid(column=0, row=0, padx=20, pady=10)
    label_end = tkinter.Label(root, text="End")
    label_end.grid(column=0, row=1, padx=20, pady=10)

    box_start = tkinter.Entry(root, width=20)
    box_start.grid(column=1, row=0, padx=20, pady=10)
    box_end = tkinter.Entry(root, width=20)
    box_end.grid(column=1, row=1, padx=20, pady=10)

    root.bind("<Return>", get)

    root.deiconify()
    root.mainloop()
    root.withdraw()

    if not box_start.get() or not box_end.get():
        msgbox("Error", "開始または終了番号が正しく入力されませんでした")
    else:
        if box_end.get() < box_start.get():
            msgbox("Error", '"Start"の値は"End"より小さくしてください')
        else:
            pass

    return box_start.get(), box_end.get()


def generateqrcode(dir, content):
    img = qrcode.make(str(content))
    img.save("{}/{}.png".format(dir, content))

    return None


def generateqrcodes(dir, start, end):
    digits = len(str(end))
    for i in range(int(start), int(end) + 1):
        content = str(i).zfill(digits)
        img = qrcode.make("{}".format(content))
        img.save("{}/{}.png".format(dir, content))

    return None


if __name__ == "__main__":
    try:
        ctypes.windll.shcore.SetProcessDpiAwareness(True)
    except:
        pass

    bookpath = openfile([("", ".xlsx")])
    sheetname = selectsheet(bookpath)

    savepath = openfolder()

    username, password = askuserinfo()
    from_email = username
    smtp_host = "smtp.office365.com"
    smtp_port = 587
    server = smtplib.SMTP(smtp_host, smtp_port)
    server.ehlo()
    server.starttls()
    server.ehlo()
    server.login(username, password)

    worksheet = openpyxl.load_workbook(bookpath)[sheetname]

    toemail_list = []
    for i in range(2, worksheet.max_row + 1):
        toemail_list.append(
            [
                worksheet.cell(row=i, column=1).value,
                worksheet.cell(row=i, column=5).value,
            ]
        )

    for i in toemail_list:
        msg = message.EmailMessage()

        msg["Subject"] = "【テスト】QRコードチケット発行 [{}]".format(i[0])
        msg["From"] = from_email
        msg["To"] = i[1]

        msg.set_content("【テスト】QRコードチケット発行\nここは本文")

        generateqrcode(savepath, str(i[0]))

        server.send_message(msg)

    server.quit()
